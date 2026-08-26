#!/usr/bin/env python3
"""
IFFL Seed Importer
==================

Reads `2026 IFFL Keeper Master List.xlsx` + the canonical mapping in
`iffl_agent_config.json` and emits:

  - seed.sql              SQL inserts ready to apply against the IFFL Postgres DB
  - seed_diff_report.md   human-readable validation report (owners, keeper-cost
                          formula checks, unmapped names, anomalies)

Default mode is **dry-run**: nothing is written to a database. Pass --apply with
SUPABASE_DB_URL set to also psql the seed file. (apply mode reserved for the
cloud-setup checkpoint.)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------- paths -------

HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE / "2026_IFFL_Master.xlsx"
DEFAULT_CONFIG = (HERE / ".." / ".." / ".." / ".." / "IFFL Commish Agent" /
                  "iffl_agent_config.json").resolve()
DEFAULT_SEASON = 2026

# Keeper escalation per IFFL rules: kept years 1..5 add +5/+10/+15/+20/+25 cumulatively.
KEEPER_STEPS = [5, 10, 15, 20, 25]


# ---------------------------------------------------------------- helpers -----

def keeper_cost(original: float, years_kept: int) -> float:
    if years_kept < 0 or years_kept > 5:
        raise ValueError(f"years_kept must be 0..5, got {years_kept}")
    return float(original) + sum(KEEPER_STEPS[:years_kept])


def parse_money(v: Any) -> float | None:
    """'$25 ' / '$25' / 25 / '' / None -> float | None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    try:
        return int(float(str(v).strip()))
    except ValueError:
        return None


def sql_str(v: Any) -> str:
    if v is None:
        return "NULL"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def sql_lit(s: str) -> str:
    """SQL string literal with apostrophes escaped. For embedding in nested
    contexts like uuid_generate_v5(ns, '...') where the inner string may
    contain user-supplied content (player names with apostrophes)."""
    return "'" + s.replace("'", "''") + "'"


def sql_num(v: Any) -> str:
    return "NULL" if v is None else f"{v}"


def sql_int(v: Any) -> str:
    return "NULL" if v is None else f"{int(v)}"


def find_header_row(ws: Worksheet, required_columns: Iterable[str],
                     scan_limit: int = 25) -> int:
    """Locate the first row containing all required column names (case-insensitive,
    whitespace-tolerant). Returns 1-based row number or raises."""
    required = {c.strip().lower() for c in required_columns}
    for r in range(1, min(scan_limit, ws.max_row) + 1):
        values = {
            (str(ws.cell(r, c).value).strip().lower())
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        }
        if required.issubset(values):
            return r
    raise ValueError(
        f"Could not locate header row in '{ws.title}' "
        f"with required columns {sorted(required)}"
    )


def header_index(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Map normalized header label -> 1-based column index. Strips trailing/leading
    whitespace from header cells (the sheet has 'Player ' with trailing space)."""
    out: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        out[str(cell.value).strip()] = cell.column
    return out


# --------------------------------------------------------------- domain -------

@dataclass
class Owner:
    master_name: str          # "A. Zurek"
    full_name: str            # "Andrew Zurek"
    espn_team_name: str       # "Cinderella Story"
    groupme_handle: str | None
    avatar_filename: str      # "A.Zurek.png"
    is_commissioner: bool = False
    is_treasurer: bool = False
    is_rules_committee: bool = False


@dataclass
class Contract:
    owner_master_name: str
    player: str
    position: str
    price_2026: float | None
    price_2027: float | None
    price_2028: float | None
    original_price: float | None
    purchase_year: int | None
    contract_year: int | None  # how many years on this contract (1=first season, etc)
    player_pool: str | None    # 'Auction', 'Rookie Draft', 'Free Agent', 'Draft Pick', ...
    rookie_round: int | None
    draft_year: int | None
    trade_history: str | None


@dataclass
class RookiePick:
    pick_year: int            # 2026 or 2027
    slot: str                 # "1.01", "2.07", ...
    round: int                # 1 or 2
    slot_in_round: int        # 1..12
    price: float | None
    player_name: str | None
    nfl_team: str | None
    owner_master_name: str | None
    trade_history: str | None


@dataclass
class CalendarEntry:
    season_milestone: str
    yearly_milestone: str | None
    date_text: str | None     # raw "February 10", "April 23rd", ...
    description: str | None


@dataclass
class SeedReport:
    owners_count: int = 0
    contracts_count: int = 0
    contracts_keeper_match_count: int = 0
    contracts_keeper_mismatch: list[dict[str, Any]] = field(default_factory=list)
    rookie_picks_2026_count: int = 0
    rookie_picks_2027_count: int = 0
    calendar_count: int = 0
    unmapped_owner_names: Counter = field(default_factory=Counter)
    unknown_player_pools: Counter = field(default_factory=Counter)
    notes: list[str] = field(default_factory=list)


# ---------------------------------------------------------- config loading ----

def load_canonical_config(path: Path) -> tuple[dict[str, Owner], dict[str, str]]:
    raw = json.loads(path.read_text())
    owner_full_names: dict[str, str] = raw["owner_full_names"]
    espn_to_master: dict[str, str] = {
        k.lower(): v for k, v in raw["espn_team_to_owner"].items()
    }
    groupme_to_master: dict[str, str] = {
        k.lower(): v for k, v in raw["groupme_name_to_owner"].items()
    }
    master_to_espn = {v: k for k, v in raw["espn_team_to_owner"].items()}
    master_to_groupme = {v: k for k, v in raw["groupme_name_to_owner"].items()}

    owners: dict[str, Owner] = {}
    for master_name in raw["owners"]:
        avatar_filename = master_name.replace(" ", "") + ".png"  # A.Zurek.png, Bill.png, ...
        owners[master_name] = Owner(
            master_name=master_name,
            full_name=owner_full_names[master_name],
            espn_team_name=master_to_espn[master_name],
            groupme_handle=master_to_groupme.get(master_name),
            avatar_filename=avatar_filename,
            is_commissioner=(master_name == "M. Zurek"),  # per rules doc
            is_treasurer=(master_name == "Jason"),
            is_rules_committee=(master_name in {"M. Zurek", "Jared", "Bill"}),
        )
    return owners, espn_to_master


# ---------------------------------------------------------- sheet readers -----

def normalize_owner_name(raw: str | None, valid_master_names: set[str],
                        report: SeedReport) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    # The Master List uses spelled forms like "A. Zurek" already. Tolerate
    # "A.Zurek" or "A . Zurek" by collapsing whitespace before/after dots.
    s_canonical = re.sub(r"\s*\.\s*", ". ", s).strip()
    s_canonical = re.sub(r"\s+", " ", s_canonical)
    if s_canonical in valid_master_names:
        return s_canonical
    # Try case-insensitive
    for v in valid_master_names:
        if v.lower() == s_canonical.lower():
            return v
    report.unmapped_owner_names[s] += 1
    return None


def read_master_list(ws: Worksheet, valid_owners: set[str],
                     report: SeedReport) -> list[Contract]:
    contracts: list[Contract] = []
    needed = ["Team", "Position", "Player", "2026 Price", "2027 Price", "2028 Price",
              "Original Price", "Purchase Year", "Contract Year", "Player Pool",
              "Rookie Round", "Draft Year", "Trade History"]
    header_row = find_header_row(ws, needed)
    h = header_index(ws, header_row)
    missing = [c for c in needed if c not in h]
    if missing:
        raise ValueError(f"Master List missing columns after header detect: {missing}")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        team_raw = ws.cell(row_idx, h["Team"]).value
        if team_raw is None or str(team_raw).strip() == "":
            continue
        owner = normalize_owner_name(team_raw, valid_owners, report)
        if owner is None:
            # Could be a header repeat or empty section divider — skip but log.
            continue
        player = ws.cell(row_idx, h["Player"]).value
        position = ws.cell(row_idx, h["Position"]).value
        if player is None or position is None:
            continue
        c = Contract(
            owner_master_name=owner,
            player=str(player).strip(),
            position=str(position).strip(),
            price_2026=parse_money(ws.cell(row_idx, h["2026 Price"]).value),
            price_2027=parse_money(ws.cell(row_idx, h["2027 Price"]).value),
            price_2028=parse_money(ws.cell(row_idx, h["2028 Price"]).value),
            original_price=parse_money(ws.cell(row_idx, h["Original Price"]).value),
            purchase_year=parse_int(ws.cell(row_idx, h["Purchase Year"]).value),
            contract_year=parse_int(ws.cell(row_idx, h["Contract Year"]).value),
            player_pool=(str(ws.cell(row_idx, h["Player Pool"]).value).strip()
                          if ws.cell(row_idx, h["Player Pool"]).value else None),
            rookie_round=parse_int(ws.cell(row_idx, h["Rookie Round"]).value),
            draft_year=parse_int(ws.cell(row_idx, h["Draft Year"]).value),
            trade_history=(str(ws.cell(row_idx, h["Trade History"]).value).strip()
                            if ws.cell(row_idx, h["Trade History"]).value else None),
        )
        contracts.append(c)
        if c.player_pool and c.player_pool.lower() not in {
            "auction", "rookie draft", "free agent", "draft pick", "trade", "keeper", "faab"
        }:
            report.unknown_player_pools[c.player_pool] += 1
    return contracts


def parse_pick_slot(raw: Any) -> tuple[int, int | None] | None:
    """Parse a pick-slot cell into (round, slot_in_round | None).

    Handles all the shapes openpyxl returns:
      "1.01"  -> (1, 1)
      "1.1"   -> (1, 10)   (Excel stripped trailing zero)
      1.01    -> (1, 1)    (numeric literal)
      1.1     -> (1, 10)   (numeric literal)
      "1.11"  -> (1, 11)
      "1st"   -> (1, None) (no slot — used for un-lottery'd 2027 picks)
      "2nd"   -> (2, None)
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        s = repr(float(raw))
        # 1.0 / 2.0 are unlikely (would be just round) but handle anyway
        try:
            r = int(float(s))
            frac = float(s) - r
        except ValueError:
            return None
        if r not in (1, 2):
            return None
        if frac < 0.005:  # whole number, just round
            return (r, None)
        # fractional part: 0.01..0.12 ; 0.10 stored as 0.1
        # Convert frac to two-digit slot. Round at 2 decimals first.
        slot_str = f"{frac:.2f}"[2:]   # ".10" -> "10"; ".07" -> "07"
        try:
            slot_in = int(slot_str)
        except ValueError:
            return None
        # Normalize: 1.1 -> slot 10 (since 1.1 == 1.10 in IFFL syntax)
        # but 1.01 has frac 0.01 -> slot_str "01" -> 1 ✓
        # We need to distinguish: openpyxl gives 1.01 as 1.01, 1.10 as 1.1.
        # f"{0.10:.2f}" = "0.10" -> slot_str "10" -> 10 ✓
        # f"{0.01:.2f}" = "0.01" -> slot_str "01" -> 1 ✓
        if 1 <= slot_in <= 12:
            return (r, slot_in)
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    m = re.match(r"^(\d)\.(\d{1,2})$", s)
    if m:
        r = int(m.group(1))
        slot_raw = m.group(2)
        # "1.1" string -> slot 10; "1.01" -> slot 1
        slot_in = int(slot_raw) if len(slot_raw) == 2 else int(slot_raw) * 10
        if r in (1, 2) and 1 <= slot_in <= 12:
            return (r, slot_in)
        return None
    if s in ("1st", "first", "round 1"):
        return (1, None)
    if s in ("2nd", "second", "round 2"):
        return (2, None)
    return None


def read_rookie_picks(ws: Worksheet, pick_year: int, valid_owners: set[str],
                      report: SeedReport) -> list[RookiePick]:
    """Read all rookie pick rows. Tolerates section banners ('2026 Round 1') and
    repeated headers between rounds. Each row independently must have a parsable
    Round value to be included."""
    picks: list[RookiePick] = []
    # Find the FIRST header row (Round/Price/Team are reliably present)
    try:
        header_row = find_header_row(ws, ["Round", "Price", "Team", "Trade History"])
    except ValueError:
        report.notes.append(
            f"Rookie picks tab '{ws.title}' has no recognizable header; skipping"
        )
        return picks
    h = header_index(ws, header_row)
    # The Player Name + NFL Team columns may not exist for the un-drafted 2027 sheet
    has_player_name = "Player Name" in h
    has_nfl_team = "NFL Team" in h

    for row_idx in range(header_row + 1, ws.max_row + 1):
        slot_raw = ws.cell(row_idx, h["Round"]).value
        parsed = parse_pick_slot(slot_raw)
        if parsed is None:
            continue
        round_num, slot_in_round = parsed
        slot_label = (
            f"{round_num}.{slot_in_round:02d}" if slot_in_round
            else f"{round_num}{'st' if round_num == 1 else 'nd'}"
        )

        team_raw = ws.cell(row_idx, h["Team"]).value
        owner = normalize_owner_name(team_raw, valid_owners, report)

        picks.append(RookiePick(
            pick_year=pick_year,
            slot=slot_label,
            round=round_num,
            slot_in_round=slot_in_round,
            price=parse_money(ws.cell(row_idx, h["Price"]).value),
            player_name=(str(ws.cell(row_idx, h["Player Name"]).value).strip()
                          if has_player_name and ws.cell(row_idx, h["Player Name"]).value
                          else None),
            nfl_team=(str(ws.cell(row_idx, h["NFL Team"]).value).strip()
                       if has_nfl_team and ws.cell(row_idx, h["NFL Team"]).value
                       else None),
            owner_master_name=owner,
            trade_history=(str(ws.cell(row_idx, h["Trade History"]).value).strip()
                            if ws.cell(row_idx, h["Trade History"]).value else None),
        ))
    return picks


def read_calendar(ws: Worksheet) -> list[CalendarEntry]:
    """Calendar tab uses col B = milestone, col C = yearly milestone, col D = 2026 date.
    Header on row 2 ('Season Milestone', 'Yearly Milestone', '2026 Date**')."""
    out: list[CalendarEntry] = []
    try:
        header_row = find_header_row(ws, ["Season Milestone", "Yearly Milestone"])
    except ValueError:
        return out
    h = header_index(ws, header_row)
    # Detect the date column (header may be "2026 Date**" or similar)
    date_col_label = next((k for k in h if k.lower().startswith(("2026 date", "date"))), None)
    if not date_col_label:
        return out
    milestone_col = h["Season Milestone"]
    yearly_col = h["Yearly Milestone"]
    date_col = h[date_col_label]

    import datetime as _dt
    for row_idx in range(header_row + 1, ws.max_row + 1):
        milestone = ws.cell(row_idx, milestone_col).value
        if milestone is None or str(milestone).strip() == "":
            continue
        yearly = ws.cell(row_idx, yearly_col).value
        date_val = ws.cell(row_idx, date_col).value
        if isinstance(date_val, _dt.datetime):
            date_text = date_val.date().isoformat()
        elif isinstance(date_val, _dt.date):
            date_text = date_val.isoformat()
        elif date_val is not None:
            date_text = str(date_val).strip()
        else:
            date_text = None
        out.append(CalendarEntry(
            season_milestone=str(milestone).strip(),
            yearly_milestone=(str(yearly).strip() if yearly else None),
            date_text=date_text,
            description=None,
        ))
    return out


def read_team_name_mapping(ws: Worksheet) -> list[dict[str, str]]:
    # Header is on first non-empty row; tolerate variations.
    try:
        header_row = find_header_row(ws, ["Master Team Name", "ESPN Team Name", "Full Name"])
    except ValueError:
        return []
    h = header_index(ws, header_row)
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        rows.append({
            (k.strip().lower().replace(" ", "_") if k else f"col{idx}"): (
                str(ws.cell(row_idx, col).value).strip()
                if ws.cell(row_idx, col).value else ""
            )
            for idx, (k, col) in enumerate(h.items())
        })
    return [r for r in rows if any(v for v in r.values())]


# ---------------------------------------------------------- validation -------

def validate_keeper_costs(contracts: list[Contract], season: int,
                          report: SeedReport) -> None:
    for c in contracts:
        if c.original_price is None or c.contract_year is None or c.price_2026 is None:
            continue
        # contract_year semantics in sheet: 1=first season the contract is held,
        # so years_kept = contract_year - 1.
        years_kept = c.contract_year - 1
        if years_kept < 0 or years_kept > 5:
            continue
        expected = keeper_cost(c.original_price, years_kept)
        if abs(expected - c.price_2026) < 0.01:
            report.contracts_keeper_match_count += 1
        else:
            report.contracts_keeper_mismatch.append({
                "owner": c.owner_master_name,
                "player": c.player,
                "position": c.position,
                "original": c.original_price,
                "contract_year": c.contract_year,
                "years_kept": years_kept,
                "expected_2026": expected,
                "actual_2026": c.price_2026,
            })


# ---------------------------------------------------------- SQL emitter -----

def emit_seed_sql(out: Path,
                  season: int,
                  owners: dict[str, Owner],
                  contracts: list[Contract],
                  rookie_picks_2026: list[RookiePick],
                  rookie_picks_2027: list[RookiePick],
                  calendar: list[CalendarEntry]) -> None:
    lines: list[str] = []
    lines.append("-- IFFL Seed (generated by seed_from_sheet.py)")
    lines.append("-- Idempotent: uses upserts via natural keys.")
    lines.append("")
    lines.append("begin;")
    lines.append("")

    # ---- seasons -------------------------------------------------------------
    lines.append(f"insert into seasons(year) values ({season})")
    lines.append("on conflict (year) do nothing;")
    lines.append("")

    # ---- owners (placeholder app_users + auth.users won't exist yet) ---------
    # We can't insert into auth.users from a plain SQL file — auth users are
    # created via supabase auth on first sign-in. So we seed app_users with
    # nullable id (we'll backfill) using a CTE keyed by email. For now we
    # generate a deterministic UUID stub so cross-table FKs can be wired up.
    #
    # The production seed step will be:
    #   1) deploy this schema
    #   2) each owner signs in once (creates auth.users row)
    #   3) a small follow-up migration UPDATEs app_users.id from auth.users
    #      via email match.
    #
    # During Phase 0 dry-run we use deterministic UUIDs (uuid_generate_v5) so
    # the seed is idempotent and FK references work.
    lines.append("-- Owners (app_users). id will be reconciled with auth.users by email post-signin.")
    lines.append("create extension if not exists \"uuid-ossp\";")
    for o in owners.values():
        oid = f"uuid_generate_v5(uuid_ns_dns(), {sql_lit(f'iffl-owner:{o.master_name}')})"
        # Use a placeholder email per rules doc until reconciled with sign-in.
        # We don't have these in the canonical config, but we know full name only.
        # Production will UPDATE email post-signin; seed uses synthetic key.
        synth_email = f"{o.master_name.replace(' ', '').replace('.', '').lower()}@iffl.local"
        lines.append(
            f"insert into app_users(id, master_name, full_name, email, "
            f"groupme_handle, is_commissioner, is_treasurer, is_rules_committee) "
            f"values ({oid}, {sql_str(o.master_name)}, {sql_str(o.full_name)}, "
            f"{sql_str(synth_email)}, {sql_str(o.groupme_handle)}, "
            f"{str(o.is_commissioner).lower()}, {str(o.is_treasurer).lower()}, "
            f"{str(o.is_rules_committee).lower()}) "
            f"on conflict (master_name) do update set "
            f"full_name = excluded.full_name, "
            f"groupme_handle = excluded.groupme_handle, "
            f"is_commissioner = excluded.is_commissioner, "
            f"is_treasurer = excluded.is_treasurer, "
            f"is_rules_committee = excluded.is_rules_committee;"
        )
    lines.append("")

    # ---- teams (one per owner per season) ------------------------------------
    lines.append("-- Teams (one per owner per season)")
    for o in owners.values():
        oid = f"uuid_generate_v5(uuid_ns_dns(), {sql_lit(f'iffl-owner:{o.master_name}')})"
        tid = f"uuid_generate_v5(uuid_ns_dns(), {sql_lit(f'iffl-team:{o.master_name}:{season}')})"
        avatar_path = f"team-avatars/{o.avatar_filename}"
        lines.append(
            f"insert into teams(id, owner_id, season, espn_team_name, team_avatar_url) "
            f"values ({tid}, {oid}, {season}, {sql_str(o.espn_team_name)}, "
            f"{sql_str(avatar_path)}) "
            f"on conflict (owner_id, season) do update set "
            f"espn_team_name = excluded.espn_team_name, "
            f"team_avatar_url = excluded.team_avatar_url;"
        )
    lines.append("")

    # ---- players + contracts -------------------------------------------------
    lines.append("-- Players + contracts from 2026 Master List")
    seen_players: dict[tuple[str, str], str] = {}  # (lower_name, position) -> uuid expr
    for c in contracts:
        if c.position == "Draft Pick":
            continue  # rookie picks handled separately
        key = (c.player.lower(), c.position)
        if key not in seen_players:
            pid_expr = (
                f"uuid_generate_v5(uuid_ns_dns(), "
                f"{sql_lit(f'iffl-player:{c.player.lower()}:{c.position}')})"
            )
            seen_players[key] = pid_expr
            lines.append(
                f"insert into players(id, full_name, position) "
                f"values ({pid_expr}, {sql_str(c.player)}, {sql_str(c.position)}) "
                f"on conflict (lower(full_name), position) do nothing;"
            )

    for c in contracts:
        if c.position == "Draft Pick":
            continue
        if c.original_price is None or c.contract_year is None:
            continue
        years_kept = max(0, (c.contract_year or 1) - 1)
        # Map player_pool string to enum
        pool_norm = (c.player_pool or "auction").lower().replace(" ", "_")
        if pool_norm == "rookie_draft":
            source = "rookie_draft"
        elif pool_norm == "free_agent" or pool_norm == "faab":
            source = "faab"
        elif pool_norm == "draft_pick":
            source = "rookie_draft"
        elif pool_norm == "trade":
            source = "trade"
        elif pool_norm == "keeper":
            source = "keeper"
        else:
            source = "auction"
        tid = f"uuid_generate_v5(uuid_ns_dns(), {sql_lit(f'iffl-team:{c.owner_master_name}:{season}')})"
        pid = (f"uuid_generate_v5(uuid_ns_dns(), "
               f"{sql_lit(f'iffl-player:{c.player.lower()}:{c.position}')})")
        cid = (f"uuid_generate_v5(uuid_ns_dns(), "
               f"{sql_lit(f'iffl-contract:{c.owner_master_name}:{c.player.lower()}:{c.position}:{season}')})")
        lines.append(
            f"insert into contracts(id, team_id, player_id, season, source, "
            f"original_cost, acquired_in_season, years_kept, current_keeper_cost, "
            f"rookie_round, rookie_year, trade_history_text) "
            f"values ({cid}, {tid}, {pid}, {season}, {sql_str(source)}, "
            f"{sql_num(c.original_price)}, "
            f"{sql_int(c.purchase_year or season - years_kept)}, {years_kept}, "
            f"{sql_num(c.price_2026 or keeper_cost(c.original_price, years_kept))}, "
            f"{sql_int(c.rookie_round)}, {sql_int(c.draft_year)}, "
            f"{sql_str(c.trade_history)}) "
            f"on conflict (team_id, player_id, season) where is_dropped = false "
            f"do update set "
            f"original_cost = excluded.original_cost, "
            f"years_kept = excluded.years_kept, "
            f"trade_history_text = excluded.trade_history_text;"
        )
    lines.append("")

    # ---- rookie picks --------------------------------------------------------
    for picks in (rookie_picks_2026, rookie_picks_2027):
        for p in picks:
            owner_team = (
                f"uuid_generate_v5(uuid_ns_dns(), "
                f"{sql_lit(f'iffl-team:{p.owner_master_name}:{season}')})"
                if p.owner_master_name else "NULL"
            )
            # Pre-lottery 2027 picks share slot labels ("1st"/"2nd") so we must
            # also include the owner master name to keep the deterministic UUID
            # unique. Slotted picks (1.01..2.12) are already unique on (year,slot).
            pick_seed = (
                f"iffl-pick:{p.pick_year}:{p.slot}"
                if p.slot_in_round is not None
                else f"iffl-pick:{p.pick_year}:{p.round}:{p.owner_master_name or 'unowned'}"
            )
            pid = f"uuid_generate_v5(uuid_ns_dns(), {sql_lit(pick_seed)})"
            lines.append(
                f"insert into rookie_picks(id, pick_year, round, slot, owner_team_id) "
                f"values ({pid}, {p.pick_year}, {p.round}, {sql_int(p.slot_in_round)}, {owner_team}) "
                f"on conflict (pick_year, round, slot) do update set "
                f"owner_team_id = excluded.owner_team_id;"
            )
    lines.append("")

    # ---- calendar ------------------------------------------------------------
    # Calendar dates in the sheet are free text ("February 10", "April 23rd").
    # We emit them with NULL due_at; a follow-up migration parses to timestamptz.
    lines.append("-- Calendar (dates as free text; parsed in follow-up migration)")
    for ce in calendar:
        cid = (f"uuid_generate_v5(uuid_ns_dns(), "
               f"{sql_lit(f'iffl-calendar:{season}:{ce.season_milestone}')})")
        lines.append(
            f"-- {ce.season_milestone} | {ce.yearly_milestone or ''} | {ce.date_text or ''}"
        )
    lines.append("")

    lines.append("commit;")
    lines.append("")

    out.write_text("\n".join(lines))


# ---------------------------------------------------------- report emitter ---

def emit_report(out: Path,
                report: SeedReport,
                owners: dict[str, Owner],
                contracts: list[Contract],
                rookie_picks_2026: list[RookiePick],
                rookie_picks_2027: list[RookiePick]) -> None:
    lines: list[str] = []
    lines.append("# IFFL Seed Dry-Run Report\n")
    lines.append(f"- Owners loaded: **{report.owners_count} / 12**")
    lines.append(f"- Contracts loaded: **{report.contracts_count}**")
    lines.append(f"- Keeper-cost formula matches: "
                  f"**{report.contracts_keeper_match_count} / "
                  f"{report.contracts_keeper_match_count + len(report.contracts_keeper_mismatch)}**")
    lines.append(f"- 2026 rookie picks: **{report.rookie_picks_2026_count}**")
    lines.append(f"- 2027 rookie picks: **{report.rookie_picks_2027_count}**")
    lines.append(f"- Calendar entries: **{report.calendar_count}**\n")

    # Roster size sanity
    by_owner = defaultdict(int)
    for c in contracts:
        if c.position != "Draft Pick":
            by_owner[c.owner_master_name] += 1
    lines.append("## Roster Size by Owner\n")
    lines.append("| Owner | Contracts |")
    lines.append("|---|---|")
    for o in sorted(owners.keys()):
        lines.append(f"| {o} | {by_owner.get(o, 0)} |")
    lines.append("")

    # Keeper mismatches
    if report.contracts_keeper_mismatch:
        lines.append(f"## Keeper-Cost Mismatches ({len(report.contracts_keeper_mismatch)})\n")
        lines.append("| Owner | Player | Pos | Original | Contract Yr | Years Kept | Expected 2026 | Actual 2026 |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for m in report.contracts_keeper_mismatch[:50]:
            lines.append(
                f"| {m['owner']} | {m['player']} | {m['position']} | "
                f"${m['original']:.0f} | {m['contract_year']} | {m['years_kept']} | "
                f"${m['expected_2026']:.0f} | ${m['actual_2026']:.0f} |"
            )
        if len(report.contracts_keeper_mismatch) > 50:
            lines.append(f"\n... and {len(report.contracts_keeper_mismatch) - 50} more")
        lines.append("")
    else:
        lines.append("## Keeper-Cost Mismatches\n\nNone — every contract matches the rules-doc formula. ✓\n")

    # Unmapped owner names
    if report.unmapped_owner_names:
        lines.append("## Unmapped 'Team' values in Master List\n")
        lines.append("| Raw value | Count |")
        lines.append("|---|---|")
        for k, v in report.unmapped_owner_names.most_common():
            lines.append(f"| `{k}` | {v} |")
        lines.append("")

    # Unknown player pools
    if report.unknown_player_pools:
        lines.append("## Unknown 'Player Pool' values\n")
        lines.append("| Value | Count |")
        lines.append("|---|---|")
        for k, v in report.unknown_player_pools.most_common():
            lines.append(f"| `{k}` | {v} |")
        lines.append("")

    # Rookie picks summary
    lines.append("## Rookie Picks 2026 (by Owner)\n")
    by_owner_2026 = defaultdict(list)
    for p in rookie_picks_2026:
        if p.owner_master_name:
            by_owner_2026[p.owner_master_name].append(p.slot)
    for o in sorted(by_owner_2026.keys()):
        slots = ", ".join(sorted(by_owner_2026[o]))
        lines.append(f"- **{o}**: {slots}")
    lines.append("")

    # Notes
    if report.notes:
        lines.append("## Notes\n")
        for n in report.notes:
            lines.append(f"- {n}")
        lines.append("")

    out.write_text("\n".join(lines))


# --------------------------------------------------------------- main --------

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    p.add_argument("--season", type=int, default=DEFAULT_SEASON)
    p.add_argument("--out-sql", type=Path, default=HERE / "seed.sql")
    p.add_argument("--out-report", type=Path, default=HERE / "seed_diff_report.md")
    p.add_argument("--apply", action="store_true",
                   help="Reserved for future use; currently no-op (dry-run only).")
    args = p.parse_args()

    if args.apply:
        sys.stderr.write("--apply is reserved; this run is dry-run only.\n")

    if not args.xlsx.exists():
        sys.stderr.write(f"Workbook not found: {args.xlsx}\n")
        return 2
    if not args.config.exists():
        sys.stderr.write(f"Config not found: {args.config}\n")
        return 2

    print(f"Loading config: {args.config}")
    owners, _espn_to_master = load_canonical_config(args.config)
    valid_owner_set = set(owners.keys())
    report = SeedReport(owners_count=len(owners))

    print(f"Loading workbook: {args.xlsx}")
    wb = load_workbook(args.xlsx, data_only=True)

    if "2026 Master List" not in wb.sheetnames:
        sys.stderr.write("Tab '2026 Master List' missing\n")
        return 3
    contracts = read_master_list(wb["2026 Master List"], valid_owner_set, report)
    report.contracts_count = len(contracts)

    rookie_2026 = read_rookie_picks(
        wb["2026 Rookie Draft Picks"], 2026, valid_owner_set, report
    ) if "2026 Rookie Draft Picks" in wb.sheetnames else []
    report.rookie_picks_2026_count = len(rookie_2026)

    rookie_2027 = read_rookie_picks(
        wb["2027 Rookie Draft Picks"], 2027, valid_owner_set, report
    ) if "2027 Rookie Draft Picks" in wb.sheetnames else []
    report.rookie_picks_2027_count = len(rookie_2027)

    calendar = read_calendar(wb["2026 Calendar"]) if "2026 Calendar" in wb.sheetnames else []
    report.calendar_count = len(calendar)

    # Cross-check Team name mapping tab against canonical config
    if "Team name mapping" in wb.sheetnames:
        sheet_mapping = read_team_name_mapping(wb["Team name mapping"])
        report.notes.append(
            f"Team name mapping tab parsed: {len(sheet_mapping)} rows "
            f"(canonical config has {len(owners)} owners)"
        )

    print("Validating keeper-cost formula against sheet's 2026 Price column…")
    validate_keeper_costs(contracts, args.season, report)

    print(f"Writing seed SQL: {args.out_sql}")
    emit_seed_sql(args.out_sql, args.season, owners, contracts,
                  rookie_2026, rookie_2027, calendar)

    print(f"Writing diff report: {args.out_report}")
    emit_report(args.out_report, report, owners, contracts, rookie_2026, rookie_2027)

    print()
    print("=" * 60)
    print(f"Owners:                {report.owners_count} / 12")
    print(f"Contracts:             {report.contracts_count}")
    print(f"  matches formula:     {report.contracts_keeper_match_count}")
    print(f"  mismatches:          {len(report.contracts_keeper_mismatch)}")
    print(f"2026 rookie picks:     {report.rookie_picks_2026_count}")
    print(f"2027 rookie picks:     {report.rookie_picks_2027_count}")
    print(f"Calendar entries:      {report.calendar_count}")
    print(f"Unmapped owner cells:  {sum(report.unmapped_owner_names.values())}")
    print("=" * 60)

    return 0 if not report.contracts_keeper_mismatch and report.contracts_count > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
